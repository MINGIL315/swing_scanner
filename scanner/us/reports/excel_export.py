"""CSV / Excel 내보내기.

``export_to_csv(scan_date)``  → ``data/exports/YYYY-MM-DD.csv``
``export_to_excel(scan_date)``→ ``data/exports/YYYY-MM-DD.xlsx`` (6 시트)
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger

from scanner.config import settings
from scanner.db.models import Universe
from scanner.db.repository import get_scan_results
from scanner.db.session import get_session

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


_COLUMNS = [
    "scan_date", "ticker", "market", "pattern_name", "confidence_score",
    "entry_price", "stop_loss", "target_price", "risk_reward_ratio",
    "trend_weekly", "passed_filters",
]

_PATTERN_LABELS = {
    "double_bottom": "쌍바닥",
    "golden_cross":  "골든크로스",
    "box_breakout":  "박스 돌파",
    "pullback":      "눌림목",
}


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------


def export_to_csv(
    scan_date: date | None = None,
    min_score: float = 0.0,
) -> Path:
    """스캔 결과를 CSV 로 내보낸다.

    Args:
        scan_date: 대상 날짜. ``None`` 이면 오늘.
        min_score: 최소 신뢰도 점수 (기본 0 = 전체).

    Returns:
        저장된 CSV 파일 경로.
    """
    if scan_date is None:
        scan_date = date.today()

    df = _build_dataframe(scan_date, min_score)
    settings.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = settings.EXPORTS_DIR / f"{scan_date.isoformat()}.csv"
    df.to_csv(out, index=False, encoding="utf-8-sig")
    logger.info("CSV 내보내기 완료: {}", out)
    return out


def export_to_excel(
    scan_date: date | None = None,
    min_score: float = 0.0,
) -> Path:
    """스캔 결과를 Excel (xlsx) 로 내보낸다.

    시트 구성:
        1. 전체결과    — 전 패턴 통합
        2. 쌍바닥
        3. 골든크로스
        4. 박스돌파
        5. 눌림목
        6. 요약통계

    Args:
        scan_date: 대상 날짜. ``None`` 이면 오늘.
        min_score: 최소 신뢰도 점수 (기본 0 = 전체).

    Returns:
        저장된 xlsx 파일 경로.
    """
    if not _OPENPYXL:
        logger.warning("openpyxl 미설치 → CSV 대체 출력")
        return export_to_csv(scan_date, min_score)

    if scan_date is None:
        scan_date = date.today()

    df_all = _build_dataframe(scan_date, min_score)
    settings.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = settings.EXPORTS_DIR / f"{scan_date.isoformat()}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # ── 시트 1: 전체 결과 ──────────────────────────────────────────
    _write_sheet(wb, "전체결과", df_all)

    # ── 시트 2~5: 패턴별 ──────────────────────────────────────────
    for pkey, plabel in _PATTERN_LABELS.items():
        sub = df_all[df_all["pattern_name"] == pkey].copy()
        _write_sheet(wb, plabel, sub)

    # ── 시트 6: 요약 통계 ──────────────────────────────────────────
    _write_summary_sheet(wb, df_all, scan_date)

    wb.save(out)
    logger.info("Excel 내보내기 완료: {}", out)
    return out


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------


def _build_dataframe(scan_date: date, min_score: float) -> pd.DataFrame:
    """DB 에서 스캔 결과를 읽어 DataFrame 을 반환한다."""
    from sqlalchemy import select

    with get_session() as session:
        rows_orm = get_scan_results(scan_date, session, min_score=min_score or None)

        tickers = list({r.ticker for r in rows_orm})
        market_map: dict[str, str] = {}
        if tickers:
            stmt = select(Universe.ticker, Universe.market).where(
                Universe.ticker.in_(tickers)
            )
            market_map = {row[0]: row[1] for row in session.execute(stmt).all()}

        records: list[dict[str, Any]] = [
            {
                "scan_date":         r.scan_date.isoformat(),
                "ticker":            r.ticker,
                "market":            market_map.get(r.ticker, "KR"),
                "pattern_name":      r.pattern_name,
                "confidence_score":  round(r.confidence_score, 2),
                "entry_price":       r.entry_price,
                "stop_loss":         r.stop_loss,
                "target_price":      r.target_price,
                "risk_reward_ratio": r.risk_reward_ratio,
                "trend_weekly":      r.trend_weekly or "",
                "passed_filters":    r.passed_filters,
            }
            for r in rows_orm
        ]

    return pd.DataFrame(records, columns=_COLUMNS) if records else pd.DataFrame(columns=_COLUMNS)


def _write_sheet(wb: "Workbook", sheet_name: str, df: pd.DataFrame) -> None:  # type: ignore[name-defined]
    """DataFrame 을 스타일이 적용된 시트로 쓴다."""
    ws = wb.create_sheet(sheet_name)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C2128")

    col_labels = {
        "scan_date":         "날짜",
        "ticker":            "종목",
        "market":            "시장",
        "pattern_name":      "패턴",
        "confidence_score":  "신뢰도",
        "entry_price":       "진입가",
        "stop_loss":         "손절가",
        "target_price":      "목표가",
        "risk_reward_ratio": "R:R",
        "trend_weekly":      "주봉추세",
        "passed_filters":    "필터통과",
    }

    # 헤더 행
    for col_idx, col_key in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_labels.get(col_key, col_key))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 데이터 행
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, col_key in enumerate(_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=getattr(row, col_key, None))

    # 신뢰도 컬럼(E=5) 색상 스케일
    score_col = get_column_letter(5)
    if len(df) > 0:
        last_row = len(df) + 1
        ws.conditional_formatting.add(
            f"{score_col}2:{score_col}{last_row}",
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="F85149",
                mid_type="num",   mid_value=70,    mid_color="FFCA28",
                end_type="num",   end_value=100,   end_color="3FB950",
            ),
        )

    # 열 너비 자동 조정
    col_widths = [12, 12, 8, 14, 10, 12, 12, 12, 8, 12, 10]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_summary_sheet(
    wb: "Workbook", df: pd.DataFrame, scan_date: date  # type: ignore[name-defined]
) -> None:
    """요약 통계 시트를 생성한다."""
    ws = wb.create_sheet("요약통계")

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1C2128")

    ws["A1"] = f"스캔 날짜: {scan_date.isoformat()}"
    ws["A1"].font = title_font
    ws["A2"] = f"총 탐지 건수: {len(df)}"

    # 패턴별 건수
    ws["A4"] = "패턴"
    ws["B4"] = "탐지 건수"
    ws["C4"] = "평균 신뢰도"
    for cell in (ws["A4"], ws["B4"], ws["C4"]):
        cell.font = header_font
        cell.fill = header_fill

    row = 5
    for pkey, plabel in _PATTERN_LABELS.items():
        sub = df[df["pattern_name"] == pkey]
        ws.cell(row=row, column=1, value=plabel)
        ws.cell(row=row, column=2, value=len(sub))
        ws.cell(
            row=row, column=3,
            value=round(sub["confidence_score"].mean(), 1) if len(sub) > 0 else None,
        )
        row += 1

    # 시장별 건수
    ws.cell(row=row + 1, column=1, value="시장")
    ws.cell(row=row + 1, column=2, value="탐지 건수")
    for cell in (ws.cell(row=row + 1, column=1), ws.cell(row=row + 1, column=2)):
        cell.font = header_font
        cell.fill = header_fill
    row += 2
    for mkt in ("KR", "US"):
        sub = df[df["market"] == mkt]
        ws.cell(row=row, column=1, value=mkt)
        ws.cell(row=row, column=2, value=len(sub))
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14


__all__ = ["export_to_csv", "export_to_excel"]
