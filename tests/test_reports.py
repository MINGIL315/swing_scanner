"""리포트 통합 테스트.

HTML 생성, CSV/Excel 내보내기 컬럼, 코멘트 생성을 검증한다.
외부 DB / 네트워크 없이 실행되도록 모든 의존성을 mock 처리한다.
"""
from __future__ import annotations

import json
from contextlib import contextmanager
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest


# ---------------------------------------------------------------------------
# 공통 픽스처
# ---------------------------------------------------------------------------

_SCAN_DATE = date(2026, 1, 15)


def _make_scan_row(
    ticker: str = "A",
    pattern: str = "pullback",
    score: float = 80.0,
) -> MagicMock:
    row = MagicMock()
    row.ticker = ticker
    row.pattern_name = pattern
    row.confidence_score = score
    row.entry_price = 100.0
    row.stop_loss = 95.0
    row.target_price = 110.0
    row.risk_reward_ratio = 2.0
    row.trend_weekly = "uptrend"
    row.passed_filters = True
    row.pattern_details = {}
    row.scan_date = _SCAN_DATE
    return row


def _make_universe_row(ticker: str, market: str = "KR") -> tuple:
    return (ticker, market)


@contextmanager
def _noop_session():
    yield MagicMock()


# ---------------------------------------------------------------------------
# HTML 리포트 테스트
# ---------------------------------------------------------------------------

class TestHtmlReport:
    def test_generate_creates_index(self, tmp_path: Path) -> None:
        """generate_daily_report 가 index.html 을 생성한다."""
        from scanner.kr.reports.html_report import generate_daily_report

        orm_rows = [_make_scan_row("A", "pullback", 80.0)]

        with (
            patch("scanner.reports.html_report.settings") as mock_settings,
            patch("scanner.reports.html_report.get_session", side_effect=_noop_session),
            patch("scanner.reports.html_report.get_scan_results", return_value=orm_rows),
            patch("scanner.reports.html_report._attach_market") as mock_attach,
            patch("scanner.reports.html_report._load_ticker_names", return_value={"A": "종목A"}),
            patch("scanner.reports.html_report._load_ohlcv", return_value=pd.DataFrame()),
            patch("scanner.reports.html_report._count_active_tickers", return_value=700),
            patch("scanner.reports.html_report.generate_comment", return_value=[
                {"label": "큰 추세", "text": "상승"},
                {"label": "패턴 분석", "text": "눌림목"},
                {"label": "진입 추천", "text": "100 진입"},
                {"label": "리스크 경고", "text": "손절 필수"},
            ]),
        ):
            mock_settings.REPORTS_DIR = tmp_path
            mock_settings_threshold = patch(
                "scanner.reports.html_report.settings_threshold", return_value=70.0
            )
            mock_settings_threshold.start()

            mock_attach.return_value = [
                {
                    "ticker": "A",
                    "pattern_name": "pullback",
                    "confidence_score": 80.0,
                    "entry_price": 100.0,
                    "stop_loss": 95.0,
                    "target_price": 110.0,
                    "risk_reward_ratio": 2.0,
                    "trend_weekly": "uptrend",
                    "passed_filters": True,
                    "pattern_details": {},
                    "market": "KR",
                }
            ]

            index_path = generate_daily_report(_SCAN_DATE)
            mock_settings_threshold.stop()

        assert index_path.exists()
        content = index_path.read_text(encoding="utf-8")
        assert "일일 스캔 리포트" in content

    def test_generate_creates_pattern_list_pages(self, tmp_path: Path) -> None:
        """4개 패턴 HTML 파일이 생성된다."""
        from scanner.kr.reports.html_report import generate_daily_report, _PATTERN_KEYS

        with (
            patch("scanner.reports.html_report.settings") as mock_settings,
            patch("scanner.reports.html_report.get_session", side_effect=_noop_session),
            patch("scanner.reports.html_report.get_scan_results", return_value=[]),
            patch("scanner.reports.html_report._attach_market", return_value=[]),
            patch("scanner.reports.html_report._load_ticker_names", return_value={}),
            patch("scanner.reports.html_report._count_active_tickers", return_value=700),
            patch("scanner.reports.html_report.settings_threshold", return_value=70.0),
        ):
            mock_settings.REPORTS_DIR = tmp_path
            generate_daily_report(_SCAN_DATE)

        report_dir = tmp_path / _SCAN_DATE.isoformat()
        for pkey in _PATTERN_KEYS:
            assert (report_dir / f"{pkey}.html").exists(), f"{pkey}.html 없음"

    def test_stock_detail_created_for_each_ticker(self, tmp_path: Path) -> None:
        """종목 수만큼 stocks/{ticker}.html 이 생성된다."""
        from scanner.kr.reports.html_report import generate_daily_report

        rows_dict = [
            {
                "ticker": "AAA",
                "pattern_name": "double_bottom",
                "confidence_score": 85.0,
                "entry_price": 50.0,
                "stop_loss": 47.0,
                "target_price": 56.0,
                "risk_reward_ratio": 2.0,
                "trend_weekly": "uptrend",
                "passed_filters": True,
                "pattern_details": {},
                "market": "US",
            },
            {
                "ticker": "BBB",
                "pattern_name": "golden_cross",
                "confidence_score": 75.0,
                "entry_price": 200.0,
                "stop_loss": 190.0,
                "target_price": 220.0,
                "risk_reward_ratio": 2.0,
                "trend_weekly": "uptrend",
                "passed_filters": True,
                "pattern_details": {},
                "market": "KR",
            },
        ]

        with (
            patch("scanner.reports.html_report.settings") as mock_settings,
            patch("scanner.reports.html_report.get_session", side_effect=_noop_session),
            patch("scanner.reports.html_report.get_scan_results", return_value=[]),
            patch("scanner.reports.html_report._attach_market", return_value=rows_dict),
            patch("scanner.reports.html_report._load_ticker_names", return_value={}),
            patch("scanner.reports.html_report._load_ohlcv", return_value=pd.DataFrame()),
            patch("scanner.reports.html_report._count_active_tickers", return_value=700),
            patch("scanner.reports.html_report.generate_comment", return_value=[]),
            patch("scanner.reports.html_report.settings_threshold", return_value=70.0),
        ):
            mock_settings.REPORTS_DIR = tmp_path
            generate_daily_report(_SCAN_DATE)

        stocks_dir = tmp_path / _SCAN_DATE.isoformat() / "stocks"
        assert (stocks_dir / "AAA.html").exists()
        assert (stocks_dir / "BBB.html").exists()


# ---------------------------------------------------------------------------
# OHLCV JSON 포맷 테스트
# ---------------------------------------------------------------------------

class TestBuildOhlcvJson:
    def test_empty_df_returns_empty_lists(self) -> None:
        from scanner.kr.reports.html_report import _build_ohlcv_json

        result = json.loads(_build_ohlcv_json(pd.DataFrame()))
        assert result["candles"] == []
        assert result["volume"] == []
        assert result["ma5"] == []
        assert result["rsi"] == []

    def test_valid_df_produces_candles(self) -> None:
        from scanner.kr.reports.html_report import _build_ohlcv_json

        df = pd.DataFrame([
            {"date": f"2026-01-{i+1:02d}", "open": 100.0, "high": 105.0,
             "low": 99.0, "close": 102.0, "volume": 1_000_000.0}
            for i in range(30)
        ])
        result = json.loads(_build_ohlcv_json(df))

        assert len(result["candles"]) == 30
        assert "time" in result["candles"][0]
        assert "open" in result["candles"][0]

    def test_volume_color_logic(self) -> None:
        from scanner.kr.reports.html_report import _build_ohlcv_json

        df = pd.DataFrame([
            {"date": "2026-01-01", "open": 100.0, "high": 105.0, "low": 99.0,
             "close": 103.0, "volume": 1_000.0},   # 양봉 → green
            {"date": "2026-01-02", "open": 105.0, "high": 106.0, "low": 100.0,
             "close": 101.0, "volume": 2_000.0},   # 음봉 → red
        ])
        result = json.loads(_build_ohlcv_json(df))

        assert result["volume"][0]["color"] == "#3fb950"
        assert result["volume"][1]["color"] == "#f85149"

    def test_ma5_starts_after_5_rows(self) -> None:
        from scanner.kr.reports.html_report import _build_ohlcv_json

        df = pd.DataFrame([
            {"date": f"2026-01-{i+1:02d}", "open": 100.0, "high": 105.0,
             "low": 99.0, "close": float(100 + i), "volume": 1_000.0}
            for i in range(20)
        ])
        result = json.loads(_build_ohlcv_json(df))

        # MA5 첫 값은 5번째 행 이후 (처음 4개는 NaN)
        assert len(result["ma5"]) <= 20
        assert len(result["ma5"]) >= 16


# ---------------------------------------------------------------------------
# CSV/Excel 내보내기 테스트
# ---------------------------------------------------------------------------

class TestExportToCsv:
    def test_csv_created(self, tmp_path: Path) -> None:
        from scanner.kr.reports.excel_export import export_to_csv

        df = pd.DataFrame([{
            "scan_date": "2026-01-15",
            "ticker": "A",
            "market": "KR",
            "pattern_name": "pullback",
            "confidence_score": 80.0,
            "entry_price": 100.0,
            "stop_loss": 95.0,
            "target_price": 110.0,
            "risk_reward_ratio": 2.0,
            "trend_weekly": "uptrend",
            "passed_filters": True,
        }])

        with (
            patch("scanner.reports.excel_export.settings") as mock_settings,
            patch("scanner.reports.excel_export.get_session", side_effect=_noop_session),
            patch("scanner.reports.excel_export._build_dataframe", return_value=df),
        ):
            mock_settings.EXPORTS_DIR = tmp_path
            out = export_to_csv(_SCAN_DATE, min_score=0.0)

        assert out.exists()
        loaded = pd.read_csv(out)
        assert "ticker" in loaded.columns
        assert "confidence_score" in loaded.columns
        assert loaded.iloc[0]["ticker"] == "A"

    def test_csv_has_all_required_columns(self, tmp_path: Path) -> None:
        from scanner.kr.reports.excel_export import export_to_csv, _COLUMNS

        empty_df = pd.DataFrame(columns=_COLUMNS)
        with (
            patch("scanner.reports.excel_export.settings") as mock_settings,
            patch("scanner.reports.excel_export.get_session", side_effect=_noop_session),
            patch("scanner.reports.excel_export._build_dataframe", return_value=empty_df),
        ):
            mock_settings.EXPORTS_DIR = tmp_path
            out = export_to_csv(_SCAN_DATE)

        loaded = pd.read_csv(out)
        for col in _COLUMNS:
            assert col in loaded.columns, f"누락 컬럼: {col}"


class TestExportToExcel:
    def test_excel_created(self, tmp_path: Path) -> None:
        from scanner.kr.reports.excel_export import export_to_excel, _COLUMNS

        df = pd.DataFrame([{
            "scan_date": "2026-01-15",
            "ticker": "B",
            "market": "US",
            "pattern_name": "double_bottom",
            "confidence_score": 85.0,
            "entry_price": 50.0,
            "stop_loss": 47.0,
            "target_price": 56.0,
            "risk_reward_ratio": 2.0,
            "trend_weekly": "uptrend",
            "passed_filters": True,
        }])

        with (
            patch("scanner.reports.excel_export.settings") as mock_settings,
            patch("scanner.reports.excel_export.get_session", side_effect=_noop_session),
            patch("scanner.reports.excel_export._build_dataframe", return_value=df),
        ):
            mock_settings.EXPORTS_DIR = tmp_path
            out = export_to_excel(_SCAN_DATE)

        assert out.suffix == ".xlsx"
        assert out.exists()

    def test_excel_has_six_sheets(self, tmp_path: Path) -> None:
        from openpyxl import load_workbook
        from scanner.kr.reports.excel_export import export_to_excel, _COLUMNS

        empty_df = pd.DataFrame(columns=_COLUMNS)
        with (
            patch("scanner.reports.excel_export.settings") as mock_settings,
            patch("scanner.reports.excel_export.get_session", side_effect=_noop_session),
            patch("scanner.reports.excel_export._build_dataframe", return_value=empty_df),
        ):
            mock_settings.EXPORTS_DIR = tmp_path
            out = export_to_excel(_SCAN_DATE)

        wb = load_workbook(out)
        assert len(wb.sheetnames) == 6
        assert "전체결과" in wb.sheetnames
        assert "요약통계" in wb.sheetnames


# ---------------------------------------------------------------------------
# CLI report / export 명령어 스모크 테스트
# ---------------------------------------------------------------------------

class TestCliReport:
    def test_report_command_success(self, tmp_path: Path) -> None:
        from typer.testing import CliRunner
        from scanner.cli import app

        runner = CliRunner()

        fake_index = tmp_path / "index.html"
        fake_index.write_text("<html/>", encoding="utf-8")

        with (
            patch("scanner.config.setup_logger"),
            patch("scanner.db.migrations.init_database"),
            patch("scanner.reports.html_report.generate_daily_report", return_value=fake_index),
        ):
            result = runner.invoke(app, ["report", "--date", "2026-01-15"])

        assert result.exit_code == 0
        assert "완료" in result.output

    def test_export_csv_command_success(self, tmp_path: Path) -> None:
        from typer.testing import CliRunner
        from scanner.cli import app

        runner = CliRunner()
        fake_csv = tmp_path / "2026-01-15.csv"
        fake_csv.write_text("a,b", encoding="utf-8")

        with (
            patch("scanner.config.setup_logger"),
            patch("scanner.db.migrations.init_database"),
            patch("scanner.reports.excel_export.export_to_csv", return_value=fake_csv),
        ):
            result = runner.invoke(app, ["export", "--format", "csv", "--date", "2026-01-15"])

        assert result.exit_code == 0
        assert "완료" in result.output

    def test_export_unknown_format_exits_nonzero(self) -> None:
        from typer.testing import CliRunner
        from scanner.cli import app

        runner = CliRunner()
        with (
            patch("scanner.config.setup_logger"),
            patch("scanner.db.migrations.init_database"),
        ):
            result = runner.invoke(app, ["export", "--format", "pdf"])

        assert result.exit_code != 0
